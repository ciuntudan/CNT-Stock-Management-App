import pandas as pd
from datetime import datetime
from database.models import Product, StockMovement
from database.operations import get_all_products, get_stock_movements
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def export_to_excel(filename, data_type='products', **kwargs):
    if data_type == 'products':
        _export_products_to_excel(filename)
    elif data_type == 'stock_movements':
        _export_stock_movements_to_excel(filename, **kwargs)
    elif data_type == 'sales_report':
        _export_sales_report_to_excel(filename, **kwargs)

def export_to_csv(filename, data_type='products', **kwargs):
    if data_type == 'products':
        _export_products_to_csv(filename)
    elif data_type == 'stock_movements':
        _export_stock_movements_to_csv(filename, **kwargs)
    elif data_type == 'sales_report':
        _export_sales_report_to_csv(filename, **kwargs)

def _export_products_to_excel(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produse"
    
    # Headers
    headers = [
        'ID',
        'Nume',
        'Categorie',
        'Furnizor',
        'Cantitate',
        'Unitate',
        'Pret unitar (RON)',
        'Pret fara TVA (RON)',
        'Dimensiuni',
        'Prag minim'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # Data
    products = get_all_products()
    for row, product in enumerate(products, 2):
        # Calculate price without VAT
        price_without_vat = product.unit_price / 1.19
        
        # Format dimensions if they exist
        dimensions = ""
        if product.width and product.height:
            if product.width < 1 and product.height < 1:
                dimensions = f"{product.width * 100:.1f} x {product.height * 100:.1f} cm"
            else:
                dimensions = f"{product.width:.2f} x {product.height:.2f} m"
        
        data = [
            product.id,
            product.name,
            product.category.name if product.category else "",
            product.supplier.name if product.supplier else "",
            product.quantity,
            product.measure_unit.value,
            f"{product.unit_price:.2f}",
            f"{price_without_vat:.2f}",
            dimensions,
            product.alert_threshold
        ]
        
        for col, value in enumerate(data, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = value
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    workbook.save(filename)

def _export_products_to_csv(filename):
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # Headers
        writer.writerow([
            'ID',
            'Nume',
            'Categorie',
            'Furnizor',
            'Cantitate',
            'Unitate',
            'Pret unitar (RON)',
            'Pret fara TVA (RON)',
            'Dimensiuni',
            'Prag minim'
        ])
        
        # Data
        products = get_all_products()
        for product in products:
            # Calculate price without VAT
            price_without_vat = product.unit_price / 1.19
            
            # Format dimensions if they exist
            dimensions = ""
            if product.width and product.height:
                if product.width < 1 and product.height < 1:
                    dimensions = f"{product.width * 100:.1f} x {product.height * 100:.1f} cm"
                else:
                    dimensions = f"{product.width:.2f} x {product.height:.2f} m"
            
            writer.writerow([
                product.id,
                product.name,
                product.category.name if product.category else "",
                product.supplier.name if product.supplier else "",
                product.quantity,
                product.measure_unit.value,
                f"{product.unit_price:.2f}",
                f"{price_without_vat:.2f}",
                dimensions,
                product.alert_threshold
            ])

def _export_stock_movements_to_excel(filename, start_date=None, end_date=None, product_id=None):
    movements = get_stock_movements(product_id, start_date, end_date)
    data = []
    for movement in movements:
        data.append({
            'Date': movement.timestamp,
            'Product': movement.product.name,
            'SKU': movement.product.sku,
            'Change': movement.quantity_changed,
            'Type': movement.reference_type,
            'Reference ID': movement.reference_id,
            'Performed By': movement.performed_by.username if movement.performed_by else '',
            'Notes': movement.notes or ''
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel writer object
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # Write main data
    df.to_excel(writer, index=False, sheet_name='Stock Movements')
    
    # Add summary sheet
    summary_data = {
        'Total Movements': len(movements),
        'Period Start': start_date or 'All Time',
        'Period End': end_date or 'Present',
        'Generated On': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    summary_df = pd.DataFrame([summary_data])
    summary_df.to_excel(writer, index=False, sheet_name='Summary')
    
    writer.close()

def _export_stock_movements_to_csv(filename, start_date=None, end_date=None, product_id=None):
    movements = get_stock_movements(product_id, start_date, end_date)
    data = []
    for movement in movements:
        data.append({
            'Date': movement.timestamp,
            'Product': movement.product.name,
            'SKU': movement.product.sku,
            'Change': movement.quantity_changed,
            'Type': movement.reference_type,
            'Reference ID': movement.reference_id,
            'Performed By': movement.performed_by.username if movement.performed_by else '',
            'Notes': movement.notes or ''
        })
    
    df = pd.DataFrame(data)
    df.to_csv(filename, index=False)

def _export_sales_report_to_excel(filename, start_date=None, end_date=None):
    from database.operations import get_session
    from database.models import SalesOrder, SalesOrderItem
    
    session = get_session()
    query = session.query(SalesOrder)
    
    if start_date:
        query = query.filter(SalesOrder.order_date >= start_date)
    if end_date:
        query = query.filter(SalesOrder.order_date <= end_date)
    
    orders = query.all()
    
    # Prepare sales data
    sales_data = []
    for order in orders:
        for item in order.items:
            sales_data.append({
                'Order ID': order.id,
                'Date': order.order_date,
                'Customer': order.customer.name,
                'Product': item.product.name,
                'SKU': item.product.sku,
                'Quantity': item.quantity,
                'Unit Price': item.unit_price,
                'Total': item.quantity * item.unit_price,
                'Status': order.status
            })
    
    # Create Excel writer object
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # Write detailed sales data
    sales_df = pd.DataFrame(sales_data)
    sales_df.to_excel(writer, index=False, sheet_name='Sales Details')
    
    # Create summary data
    summary_data = {
        'Total Orders': len(orders),
        'Total Revenue': sum(order.total_amount for order in orders),
        'Period Start': start_date or 'All Time',
        'Period End': end_date or 'Present',
        'Generated On': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    summary_df = pd.DataFrame([summary_data])
    summary_df.to_excel(writer, index=False, sheet_name='Summary')
    
    # Add product summary
    product_summary = pd.DataFrame(sales_data).groupby('Product').agg({
        'Quantity': 'sum',
        'Total': 'sum'
    }).reset_index()
    product_summary.to_excel(writer, index=False, sheet_name='Product Summary')
    
    writer.close()
    session.close()

def _export_sales_report_to_csv(filename, start_date=None, end_date=None):
    from database.operations import get_session
    from database.models import SalesOrder, SalesOrderItem
    
    session = get_session()
    query = session.query(SalesOrder)
    
    if start_date:
        query = query.filter(SalesOrder.order_date >= start_date)
    if end_date:
        query = query.filter(SalesOrder.order_date <= end_date)
    
    orders = query.all()
    
    # Prepare sales data
    sales_data = []
    for order in orders:
        for item in order.items:
            sales_data.append({
                'Order ID': order.id,
                'Date': order.order_date,
                'Customer': order.customer.name,
                'Product': item.product.name,
                'SKU': item.product.sku,
                'Quantity': item.quantity,
                'Unit Price': item.unit_price,
                'Total': item.quantity * item.unit_price,
                'Status': order.status
            })
    
    df = pd.DataFrame(sales_data)
    df.to_csv(filename, index=False)
    
    session.close() 